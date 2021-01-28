from openpyxl import Workbook
from openpyxl import load_workbook
class Produs:
    def __init__(self, full_name, nume, pret, detinut, tip):
        self.tip = tip
        self.full_name = full_name
        self.nume = nume
        self.pret = []
        self.pret.append(float(pret))
        self.detinut = detinut
        self.tendinta = 0
        try:
          self.pret_initial = pret
          self.pret_max = pret
          self.pret_min = pret
        except:
            self.pret_initial = 0
            self.pret_max = 0
            self.pret_min = 999999999


    def modif_pret(self, pret):
        if float(pret) > self.pret[0]:
            self.creste()
        else:
            self.scade()
        if pret > self.pret_max:
            self.pret_max = pret
        if pret < self.pret_max:
            self.pret_min = pret
        self.pret.insert(0, pret)
        if len(self.pret) > 500:
            self.pret.pop()


    def cumpar(self):
        self.detinut = True


    def vandut(self):
        self.detinut = False


    def status(self):
        return self.detinut


    def creste(self):
        self.tendinta = 1


    def scade(self):
        self.tendinta = -1


    def afisare_produs(self):
        print(f'{self.tip} - {self.nume} a avut pretul initial de {self.pret_initial} si are tendinta {self.tendinta} si are preturile : ')
        for pret1 in self.pret:
            print(pret1)

def produs_to_excel(lista_prod,filename ):
    workbook = Workbook()
    sheet = workbook.active
    sheet.row = 1
    sheet.column = 1
    for prod in lista_prod:
        sheet.column = 1
        sheet.cell(sheet.row, sheet.column).value = prod.tip
        sheet.column += 1
        sheet.cell(sheet.row, sheet.column).value = prod.full_name
        sheet.column += 1
        sheet.cell(sheet.row, sheet.column).value = prod.nume
        sheet.column += 1
        sheet.cell(sheet.row, sheet.column).value = prod.pret_initial
        sheet.column += 1
        sheet.cell(sheet.row, sheet.column).value = prod.pret_max
        sheet.column += 1
        sheet.cell(sheet.row, sheet.column).value = prod.pret_min
        sheet.column += 1
        sheet.cell(sheet.row, sheet.column).value = prod.detinut
        sheet.column += 1
        for pret in prod.pret:
            sheet.cell(sheet.row, sheet.column).value = float(pret)
            sheet.column += 1

        sheet.row += 1
    workbook.save(filename=filename)
    print('end to excel')

def readFromExcel(filename, lista_produse):
    workbook = load_workbook(filename = filename)
    sheet = workbook.active
    rand = 1
    for row in sheet.iter_rows(values_only=True):
        column = 1
        tip = sheet.cell(rand, column).value
        column += 1
        full_name = sheet.cell(rand, column).value
        column += 1
        nume = sheet.cell(rand, column).value
        column += 1
        pret_initial = sheet.cell(rand, column).value
        column += 1
        pret_max = sheet.cell(rand, column).value
        column += 1
        pret_min = sheet.cell(rand, column).value
        column += 1
        detinut = sheet.cell(rand, column).value
        column += 1
        prod = Produs(full_name, nume, pret_initial, detinut, tip)
        prod.pret_initial = pret_initial
        prod.pret_max = pret_max
        prod.pret_min = pret_min

        for i in range(1, 50):
        #    print(i)
            if sheet.cell(rand, column).value == None:
                break
            else:
            #    print("de ce "+ str(sheet.cell(rand, column).value))
                prod.modif_pret(sheet.cell(rand, column).value)
                column += 1
        rand += 1
        lista_produse.append(prod)




